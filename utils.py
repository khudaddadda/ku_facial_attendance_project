import os
from datetime import datetime
from PIL import Image
import customtkinter as ctk
from config import IMAGES_PATH, ENCODINGS_PATH, REPORTS_PATH, COLORS


def ensure_directories():
    """Create necessary directories if they don't exist"""
    directories = [IMAGES_PATH, ENCODINGS_PATH, REPORTS_PATH]
    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f" Created directory: {directory}")


def get_current_date():
    """Get current date in YYYY-MM-DD format"""
    return datetime.now().strftime("%Y-%m-%d")


def get_current_time():
    """Get current time in HH:MM:SS format"""
    return datetime.now().strftime("%H:%M:%S")


def get_current_datetime():
    """Get current datetime formatted"""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_display_date():
    """Get display-friendly date format"""
    return datetime.now().strftime("%A, %B %d, %Y")


def get_display_time():
    """Get display-friendly time format"""
    return datetime.now().strftime("%I:%M:%S %p")


def format_date(date_str):
    """Format date string to display format"""
    try:
        date_obj = datetime.strptime(str(date_str), "%Y-%m-%d")
        return date_obj.strftime("%b %d, %Y")
    except:
        return date_str


def format_time(time_str):
    """Format time string to display format"""
    try:
        time_obj = datetime.strptime(str(time_str), "%H:%M:%S")
        return time_obj.strftime("%I:%M %p")
    except:
        return time_str


def generate_student_id():
    """Generate unique student ID"""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"STU{timestamp}"


def save_image(image_array, filename):
    """Save image array to file"""
    try:
        ensure_directories()
        filepath = os.path.join(IMAGES_PATH, filename)
        Image.fromarray(image_array).save(filepath)
        return filepath
    except Exception as e:
        print(f" Error saving image: {e}")
        return None


def resize_image(image_path, size=(200, 200)):
    """Resize image to specified size"""
    try:
        img = Image.open(image_path)
        img = img.resize(size, Image.Resampling.LANCZOS)
        return img
    except Exception as e:
        print(f" Error resizing image: {e}")
        return None


def show_toast(parent, message, message_type="success"):
    """Show toast notification"""
    colors = {
        "success": "#10b981",
        "error": "#ef4444",
        "info": "#06b6d4",
        "warning": "#f59e0b"
    }
    
    color = colors.get(message_type, colors["info"])
    
    # Create toast frame
    toast = ctk.CTkFrame(parent, fg_color=color, corner_radius=10)
    toast.place(relx=0.5, rely=0.1, anchor="center")
    
    # Toast label - plain text only (no icon)
    label = ctk.CTkLabel(
        toast,
        text=message,  # Just the message, no icon
        font=("Roboto", 14, "bold"),
        text_color="white",
        padx=20,
        pady=10
    )
    label.pack()
    
    # Auto-hide after 3 seconds
    parent.after(3000, lambda: toast.destroy())


def validate_email(email):
    """Validate email format"""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_phone(phone):
    """Validate phone number format (10 digits)"""
    import re
    pattern = r'^[0-9]{10}$'
    return re.match(pattern, phone) is not None


def center_window(window, width, height):
    """Center window on screen"""
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")


def clear_frame(frame):
    """Clear all widgets from frame"""
    for widget in frame.winfo_children():
        widget.destroy()


def export_to_excel(data, filename, headers):
    """Export data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        ensure_directories()
        filepath = os.path.join(REPORTS_PATH, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header style
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return filepath
    except Exception as e:
        print(f" Error exporting to Excel: {e}")
        return None


def get_responsive_font_size(base_size=14, window_width=None):
    """Get font size based on window width"""
    if window_width is None:
        try:
            import tkinter as tk
            root = tk._default_root
            if root:
                window_width = root.winfo_width()
        except:
            window_width = 1200
    
    scale = min(max(window_width / 1200, 0.7), 1.3)
    return int(base_size * scale)


def get_responsive_padding(base_padding=10, window_width=None):
    """Get padding based on window width"""
    if window_width is None:
        try:
            import tkinter as tk
            root = tk._default_root
            if root:
                window_width = root.winfo_width()
        except:
            window_width = 1200
    
    scale = min(max(window_width / 1200, 0.7), 1.3)
    return int(base_padding * scale)


class LoadingDialog(ctk.CTkToplevel):
    """Loading dialog with progress indication"""

    def __init__(self, parent, title="Processing", message="Please wait..."):
        super().__init__(parent)

        self.title(title)
        self.geometry("400x150")
        self.resizable(False, False)

        # Center on parent
        self.transient(parent)
        self.grab_set()

        # Main frame
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Message label
        self.message_label = ctk.CTkLabel(
            main_frame,
            text=message,
            font=("Roboto", 16),
            wraplength=350
        )
        self.message_label.pack(pady=(0, 15))

        # Progress bar
        self.progress = ctk.CTkProgressBar(
            main_frame,
            width=360,
            height=10,
            mode="indeterminate"
        )
        self.progress.pack()
        self.progress.start()

        # Center on screen
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def update_message(self, message):
        """Update loading message"""
        self.message_label.configure(text=message)
        self.update()

    def close(self):
        """Close loading dialog"""
        self.progress.stop()
        self.destroy()